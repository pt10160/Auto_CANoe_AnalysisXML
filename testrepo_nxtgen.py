#xml读取导出
#Version 1.0 alpha 
#Author: Martin Li
import os
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np

red_fill = PatternFill(start_color="ED583A", end_color="ED583A", fill_type="solid")
title_fill = PatternFill(start_color="B3B3B3", end_color="B3B3B3", fill_type="solid")
subtitle_fill = PatternFill(start_color="729E84", end_color="729E84", fill_type="solid")
bold_font = Font(bold=True)
white_font = Font(color="FFFFFF")


class TestRepo:
    def __init__(self):
        pass

    def parse_xml_file(self, file_path):
            tree = ET.parse(file_path)
            root = tree.getroot()

            # 在这里根据需要处理XML文件的内容
            # 可以使用root.findall()、root.find()等方法来获取特定元素

            # 示例：获取所有testgroup元素
            testgroups = root.findall('testgroup')
            result_rows = []  # 存储结果的行列表

            for testgroup in testgroups:
                title = testgroup.find('title').text
                result_rows.append([f"{title}"])  # 添加testgroup标题为单独的行

                testcases = testgroup.findall('testcase')
                count = 0  # 计数器变量
                pass_count = 0  # pass计数器
                fail_count = 0  # fail计数器

                for testcase in testcases:
                    title = testcase.find('title').text

                    # 检查teststep、verdict和checkstatistic中是否有fail
                    has_fail = False
                    teststeps = testcase.findall('.//teststep')
                    for teststep in teststeps:
                        result = teststep.attrib.get('result', 'pass')
                        if result == 'fail':
                            has_fail = True
                            break

                    verdict_element = testcase.find('verdict')
                    verdict_result = verdict_element.attrib.get('result', 'pass') if verdict_element is not None else 'pass'

                    checkstatistic_element = testcase.find('.//checkstatistic')
                    checkstatistic_result = checkstatistic_element.attrib.get('result', 'pass') if checkstatistic_element is not None else 'pass'

                    failure_ratio = None  # 存储 Failure ratio 的值

                    if not has_fail and verdict_result == 'pass' and checkstatistic_result == 'pass':
                        result_rows.append([title, "Pass"])
                        #pass_count += 1
                    else:


                        # 查找 Failure ratio 的值
                        if checkstatistic_element is not None:
                            xinfo_elements = checkstatistic_element.findall('.//xinfo')
                            for xinfo_element in xinfo_elements:
                                name_element = xinfo_element.find('name')
                                if name_element is not None and name_element.text == "Failure ratio (in %)":
                                    description_element = xinfo_element.find('description')
                                    if description_element is not None:
                                        failure_ratio = description_element.text
                        failute_ratiotic = failure_ratio.split("%")[0]                
                        result_rows.append([title, "Fail",failute_ratiotic])
                        #fail_count += 1

                    #count += 1
 
                    # 将 Failure ratio 添加到失败行后面
                    #if failure_ratio is not None:
                    #   result_rows.append(["", "", "", failure_ratio])
            data = [result_rows]


            for sublist in data:
                for item in sublist[1:]:
                    if len(item) >= 2 and item[1] == "Pass":
                        pass_count += 1
                        sublist.remove(item)
                    elif len(item) >= 2 and item[1] == "Fail":
                        fail_count += 1


            print("Pass count:", pass_count)
            print("Fail count:", fail_count)

            return result_rows, pass_count, fail_count

    
    def get_summary_counts(pass_count, fail_count):
        # Calculate total, pass, and fail counts

        total_count = pass_count + fail_count
        return total_count, pass_count, fail_count

    def write_summary(self, ws, total_count, pass_count, fail_count):
        # Write the summary to the worksheet
        ws.append(["Total Test Case Titles:", total_count])
        ws.append(["Pass Count:", pass_count])
        ws.append(["Fail Count:", fail_count])
        ws.append(["\n"])

    def adjust_column_widths(self, ws):
        # Adjust column widths to fit content
        for column in ws.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            column_letter = get_column_letter(column[0].column)
            ws.column_dimensions[column_letter].width = adjusted_width

    def create_filtered_sheet(self, wb, ws):
        # Create a new filtered sheet and copy non-Pass rows
        ws2 = wb.create_sheet(title="Filtered")
        for row in ws.iter_rows():
            if len(row) > 1 and row[1].value != "Pass":
                ws2.append([cell.value for cell in row])
                for new_cell, old_cell in zip(ws2[ws2.max_row], row):
                    if old_cell.has_style:
                        new_cell.font = old_cell.font.copy()
                        new_cell.border = old_cell.border.copy()
                        new_cell.fill = old_cell.fill.copy()
                        new_cell.number_format = old_cell.number_format
                        new_cell.protection = old_cell.protection.copy()
                        new_cell.alignment = old_cell.alignment.copy()
        return ws2


    # In the TestRepo class, modify the process_xml_files method as follows
    def process_xml_files(self, folder_path):
        output_file = "test_results.xlsx"
        # Check if the output file already exists
        if os.path.exists(output_file):
            # Load the existing workbook
            wb = load_workbook(output_file)
            # Use the existing worksheet
            ws = wb.active
        else:
            # If the output file does not exist, create a new workbook and worksheet
            wb = Workbook()
            ws = wb.active



        for filename in os.listdir(folder_path):
            if filename.endswith(".xml"):
                file_path = os.path.join(folder_path, filename)
                group_name = filename.replace("_test_report.xml", "")
                ws.append([])
                ws.append([f"{group_name}"])
                ws.cell(row=ws.max_row, column=1).font = bold_font
                ws.cell(row=ws.max_row, column=1).fill = title_fill

                result_rows, pass_count, fail_count = self.parse_xml_file(file_path)

                for row in result_rows:
                    ws.append(row)
                    if len(row) > 1 and row[1] == "Fail":
                        ws.cell(row=ws.max_row, column=1).fill = red_fill
                        ws.cell(row=ws.max_row, column=1).font = white_font
                    elif len(row) > 0 and row[0].startswith("XH8"):
                        ws.cell(row=ws.max_row, column=1).fill = subtitle_fill
                        ws.cell(row=ws.max_row, column=1).font = bold_font

                        

                total_count, pass_count, fail_count = TestRepo.get_summary_counts(pass_count, fail_count)
                self.write_summary(ws, total_count, pass_count, fail_count)
                ws.append([])

        self.adjust_column_widths(ws)

        # Save the updated workbook to the output file
        wb.save(output_file)

def find_row_number(file_path):
    # 读取Excel文件
    df = pd.read_excel(file_path)
    
    # 筛选满足条件的行
    condition = df.iloc[:, 0] == 'ADCPublic'
    filtered_rows = df[condition]
    
    # 检查筛选结果的长度
    if len(filtered_rows) < 2:
        print("没有足够的满足条件的行")
        return None
    
    # 获取第二个满足条件的行的行数
    row_number = filtered_rows.index[1] + 1
    
    return row_number


def move_table(file_path, start_row):
    
    if start_row is None:
        return

    # 读取Excel文件
    df = pd.read_excel(file_path)
    
    # 获取表格的最大列数
    max_column = df.shape[1]
    
    # 读取指定行之后的每行的前三个单元格内容
    table_data = df.iloc[start_row-1:, :3]
    
    # 在目标位置添加新列
    for i in range(3):
        # Find the first column name that does not exist
        counter = i
        new_column_name = f'New Column{counter}'
        while new_column_name in df.columns:
            counter += 1
            new_column_name = f'New Column{counter}'

        df.insert(loc=max_column + i, column=new_column_name, value=pd.NA)

    # 将小表格平移到大表格的新列
    df.iloc[:table_data.shape[0], max_column:max_column+3] = table_data.values

    # Clear original data
    df.iloc[start_row-1:, :3] = np.nan
    
    df.to_excel(file_path, index=False)


def adjust_excel_column_widths(file_path):
    # Load the workbook
    wb = load_workbook(file_path)

    # Loop over every sheet in the workbook
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # Save the workbook to a new file
    wb.save(file_path)




if __name__ == "__main__":
    test_repo = TestRepo()
    
    folder_path = r"C:\\Users\\ROG\\Desktop\\autorepo"
    file_path = "test_results.xlsx"
    test_repo.process_xml_files(folder_path)
    row_number = find_row_number(file_path)
    print(row_number)

    move_table(file_path,row_number)
    adjust_excel_column_widths(file_path)










