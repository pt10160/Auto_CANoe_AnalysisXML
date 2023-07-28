import os
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

class TestRepo:
    def __init__(self):
        pass

    def parse_xml_file(self, file_path):
        tree = ET.parse(file_path)
        root = tree.getroot()

        # 在这里根据需要处理XML文件的内容
        # 可以使用root.findall()、root.find()等方法来获取特定元素

        # 示例：获取所有testcase元素
        testcases = root.findall('.//testcase')
        count = 0  # 计数器变量
        pass_count = 0  # pass计数器
        fail_count = 0  # fail计数器
        result_rows = []  # 存储结果的行列表
        for testcase in testcases:
            title = testcase.find('title').text

            # 检查teststep、verdict和checkstatistic中是否有fail
            has_fail = False
            teststeps = testcase.findall('.//teststep')
            for teststep in teststeps:
                result = teststep.attrib.get('result', 'pass')
                if result == 'fail':
                    has_fail = True
                    break

            verdict_element = testcase.find('verdict')
            verdict_result = verdict_element.attrib.get('result', 'pass') if verdict_element is not None else 'pass'

            checkstatistic_element = testcase.find('.//checkstatistic')
            checkstatistic_result = checkstatistic_element.attrib.get('result', 'pass') if checkstatistic_element is not None else 'pass'

            if not has_fail and verdict_result == 'pass' and checkstatistic_result == 'pass':
                result_rows.append([title, "Pass"])
                pass_count += 1
            else:
                result_rows.append([title, "Fail"])
                fail_count += 1

            count += 1

        result_rows.append(["Total Test Case Titles:", count])
        result_rows.append(["Pass Count:", pass_count])
        result_rows.append(["Fail Count:", fail_count])

        return result_rows

if __name__ == "__main__":
    test_repo = TestRepo()
    folder_path = r"C:\\Users\\ROG\\Desktop\\autorepo"
    output_file = "test_results.xlsx"

    wb = Workbook()
    ws = wb.active

    # 创建红色填充样式、绿色填充样式和加粗字体样式
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_fill = PatternFill(start_color="19A558", end_color="19A558", fill_type="solid")
    title_fill = PatternFill(start_color="B3B3B3", end_color="B3B3B3", fill_type="solid")
    item_fill = PatternFill(start_color="E5E5E5", end_color="E5E5E5", fill_type="solid")
    bold_font = Font(bold=True)

    for filename in os.listdir(folder_path):
        if filename.endswith(".xml"):
            file_path = os.path.join(folder_path, filename)
            group_name = filename.replace("_test_report.xml", "")             
            ws.append([f"File: {group_name}"])
            ws.cell(row=ws.max_row, column=1).font = bold_font
            result_rows = test_repo.parse_xml_file(file_path)
            for row in result_rows:
                ws.append(row)
                if row[-1] == "Fail":
                    ws.cell(row=ws.max_row, column=len(row)).fill = red_fill
                    ws.cell(row=ws.max_row, column=1).fill = title_fill  # 应用标题填充样式
                elif row[-1] == "Pass":
                    ws.cell(row=ws.max_row, column=len(row)).fill = green_fill
                    ws.cell(row=ws.max_row, column=1).fill = item_fill  # 应用小项填充样式
            ws.append([])


    # 调整列宽度以适应内容
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_file)
