import os
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

class TestRepo:
    def __init__(self):
        pass

    def parse_xml_file(self, file_path):
        tree = ET.parse(file_path)
        root = tree.getroot()

        # 在这里根据需要处理XML文件的内容
        # 可以使用root.findall()、root.find()等方法来获取特定元素

        # 示例：获取所有testgroup元素
        testgroups = root.findall('testgroup')
        result_rows = []  # 存储结果的行列表

        for testgroup in testgroups:
            title = testgroup.find('title').text
            result_rows.append([f"{title}"])  # 添加testgroup标题为单独的行

            testcases = testgroup.findall('testcase')
            count = 0  # 计数器变量
            pass_count = 0  # pass计数器
            fail_count = 0  # fail计数器

            for testcase in testcases:
                title = testcase.find('title').text

                # 检查teststep、verdict和checkstatistic中是否有fail
                has_fail = False
                teststeps = testcase.findall('.//teststep')
                for teststep in teststeps:
                    result = teststep.attrib.get('result', 'pass')
                    if result == 'fail':
                        has_fail = True
                        break

                verdict_element = testcase.find('verdict')
                verdict_result = verdict_element.attrib.get('result', 'pass') if verdict_element is not None else 'pass'

                checkstatistic_element = testcase.find('.//checkstatistic')
                checkstatistic_result = checkstatistic_element.attrib.get('result', 'pass') if checkstatistic_element is not None else 'pass'

                failure_ratio = None  # 存储 Failure ratio 的值

                if not has_fail and verdict_result == 'pass' and checkstatistic_result == 'pass':
                    result_rows.append([title, "Pass"])
                    pass_count += 1
                else:


                    # 查找 Failure ratio 的值
                    if checkstatistic_element is not None:
                        xinfo_elements = checkstatistic_element.findall('.//xinfo')
                        for xinfo_element in xinfo_elements:
                            name_element = xinfo_element.find('name')
                            if name_element is not None and name_element.text == "Failure ratio (in %)":
                                description_element = xinfo_element.find('description')
                                if description_element is not None:
                                    failure_ratio = description_element.text
                    failute_ratiotic = failure_ratio.split("%")[0]                
                    result_rows.append([title, "Fail",failute_ratiotic])
                    fail_count += 1

                count += 1

                # 将 Failure ratio 添加到失败行后面
                #if failure_ratio is not None:
                 #   result_rows.append(["", "", "", failure_ratio])

        return result_rows

if __name__ == "__main__":
    test_repo = TestRepo()
    folder_path = r"C:\\Users\\ROG\\Desktop\\autorepo"  #文件夹地址
    output_file = "test_results.xlsx"                   #输出文件名

    wb = Workbook()
    ws = wb.active

    # 创建填充样式和加粗字体样式
    red_fill = PatternFill(start_color="ED583A", end_color="ED583A", fill_type="solid")
    green_fill = PatternFill(start_color="19A558", end_color="19A558", fill_type="solid")
    title_fill = PatternFill(start_color="B3B3B3", end_color="B3B3B3", fill_type="solid")
    subtitle_fill = PatternFill(start_color="729E84", end_color="729E84", fill_type="solid")
    item_fill = PatternFill(start_color="E5E5E5", end_color="E5E5E5", fill_type="solid")
    bold_font = Font(bold=True)
    blue_font = Font(color="0000FF")
    white_font = Font(color="FFFFFF")
    bluold_font = Font(color="B3B3B3", bold=True)

    # 遍历文件夹中的所有XML文件
    for filename in os.listdir(folder_path):
        if filename.endswith(".xml"):  # 只处理XML文件
            file_path = os.path.join(folder_path, filename)
            group_name = filename.replace("_test_report.xml", "")
            ws.append([f"{group_name}"])
            ws.cell(row=ws.max_row, column=1).font = bold_font
            result_rows = test_repo.parse_xml_file(file_path)
            ws.cell(row=ws.max_row, column=1).fill = title_fill

            for row in result_rows:
                ws.append(row)
                if len(row) > 1 and row[-2] == "Fail":
                    ws.cell(row=ws.max_row, column=1).fill = red_fill
                    #ws.cell(row=ws.max_row, column=1).fill = item_fill  # 应用标题填充样式
                    ws.cell(row=ws.max_row, column=1).font = white_font
                elif len(row) > 0 and row[-1] == "Pass":
                    ws.cell(row=ws.max_row, column=1).fill = green_fill
                    #ws.cell(row=ws.max_row, column=1).fill = item_fill  # 应用小项填充样式
                    #ws.cell(row=ws.max_row, column=1).font = blue_font
                elif len(row) > 0 and row[0].startswith("XH8"):
                    ws.cell(row=ws.max_row, column=1).fill = subtitle_fill
                    ws.cell(row=ws.max_row, column=1).font = bold_font
                    #ws.cell(row=ws.max_row, column=1).font = bluold_font

            # 统计总数并添加相应的行
            total_count = len(result_rows) - 4  # 减去标题行和统计行的数量
            pass_count = sum(1 for row in result_rows if len(row) > 0 and row[-1] == "Pass")
            fail_count = sum(1 for row in result_rows if len(row) > 0 and row[-1] == "Fail")
            ws.append(["Total Test Case Titles:", total_count])
            ws.append(["Pass Count:", pass_count])
            ws.append(["Fail Count:", fail_count])
            ws.append([])

    # 调整列宽度以适应内容
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

# 创建新工作表，将不是 "Pass" 的行复制到新工作表

    ws2 = wb.create_sheet(title="Filtered")
    for row in ws.iter_rows():
        if len(row) > 1 and row[1].value != "Pass":
            ws2.append([cell.value for cell in row])  # 复制单元格的值

            # 复制单元格的样式
            #for new_cell, old_cell in zip(ws2[-1], row):
            for new_cell, old_cell in zip(ws2[ws2.max_row], row):

                if old_cell.has_style:
                    new_cell.font = old_cell.font.copy()
                    new_cell.border = old_cell.border.copy()
                    new_cell.fill = old_cell.fill.copy()
                    new_cell.number_format = old_cell.number_format
                    new_cell.protection = old_cell.protection.copy()
                    new_cell.alignment = old_cell.alignment.copy()

    # 删除旧工作表
    wb.remove(ws)

    # 调整新工作表的列宽度以适应内容
    for column in ws2.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws2.column_dimensions[column_letter].width = adjusted_width

    # 保存工作簿到文件
    wb.save(output_file)


